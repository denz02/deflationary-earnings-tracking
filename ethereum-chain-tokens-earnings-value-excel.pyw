# Tip jar address if you appreciate this script my safemoon address 0x32D2da7308516e46D38EC638FE2144dE2250B7D3
# or Ethereum Address 0x8b3192f5eebd8579568a2ed41e6feb402f93f73f or Bitcoin address bc1q0c7zk7hgs4xzm97p07gyv2vcvsgysklw3gdrp3 Thank You 
#
#Big thanks again to Sam Brimhall
#After installing python you also need to install module requests and pandas to do this simple type in command prompt 
# following command 'python -m pip install requests' and python -m pip install pandas
########################## INSERT YOUR INFO AS NEEDED BELOW ################################################

#Contract Address - the token you wish to read value from on coingecko or https://etherscan.io (set to saitama-inu, DO NOT CHANGE unless you want different token)
contractAddr = '0x8b3192f5eebd8579568a2ed41e6feb402f93f73f'

#Wallet Address - INSERT YOUR PUBLIC WALLET ADDRESS HERE
walletAddr = ''

#API Key for etherscan - MAKE AND ACCOUNT https://etherscan.io verify your account then log back in to https://etherscan.io and on the left hand side API-KEY menu and generate a new API key. INSERT THAT API KEY HERE
ethAPIkey = ''

#Coingecko doesn't use contract address in their API they use the name if you go to coingecko and on the righ you should see
#API id and it has the name of token in this caluclation we are using saitama inu and the id is saitama-inu
tokenname = 'saitama-inu'

#Coingeko converts to currency of your choosing
tokencurr='aud'

#Location and file name example below
ethfile = 'C:/crypto/Ethereum-Daily-Tracker.xlsx'

#How the date gets recorded in Australia it is dd/mm/yyyy
datefor = "%d/%m/%Y"

#To get rid of an error it needs to have a start balance this figure is just used as movement between second row please use just whole numbers
# eg 1,000.34 please enter 1000 different regions type thousand separator differently and it causing issues 
startbal = '1'

#To get net sale figure default 12% slipage is used 
ethslipageper = '5'

#Name of the worksheet
sfworksheet = 'Daily Earnings Tracker'

#If you want to change the column header name
etheader1 = 'Date'
etheader2 = 'Opening Balance'
etheader3 = 'Daily Gain'
etheader4 = 'Closing Balance'
etheader5 = '% Increase'
etheader6 = 'Price'
etheader7 = 'Earned Reflactions'
etheader8 = 'Price From'
etheader9 = 'Gross Value'
etheader10 = 'Net Value'

# if you want to change text
ethtextp = 'Coingecko Price'
############################################################################################################

###PROGRAM CODE BELOW - DO NOT TOUCH UNLESS YOU KNOW WHAT YOU ARE DOING###
import requests
import json
import time
import csv
import pandas as pd
import os
from datetime import date
from pandas import DataFrame
import locale
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

locale.setlocale(locale.LC_ALL, '')
col_names = [etheader1,
             etheader2,
             etheader3,
             etheader4,
             etheader5,
             etheader6,
             etheader7,
             etheader8,
             etheader9,
             etheader10,]

import os.path
#If command to check does the daily tracker file exists if it doesn't than in else create it and write
#colum headers
if os.path.isfile(ethfile):
    print ("File exist")
else:
    print ("File not exist")
    df = pd.DataFrame([col_names])
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(ethfile, engine='xlsxwriter')
# Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=sfworksheet, index=False, header=False)  
# Close the Pandas Excel writer and output the Excel file.
    writer.save()
# open the workbook created to append a opening balance as engine xlsxwriter doesn't support appending rows
workbook_name = ethfile
wb = load_workbook(workbook_name)
ws = wb[sfworksheet]
if ws.cell(row = 2, column=4).value == None:
    ws['D2'] = float(startbal)
    print("New Excel Spreadsheet")
    style_1 = Font(bold=True)
    for cell in ws["1:1"]:
        cell.font = style_1
    for col_range in range(1, 11):
        cell_title = ws.cell(1, col_range)
        cell_title.fill = PatternFill(start_color="00A79D", end_color="00A79D", fill_type="solid")
    for col_range in range(1, 11):
        cell_title = ws.cell(1, col_range)
        cell.font = style_1
else:
    print("Existing Worksheet")
#Save the spredsheet  
wb.save(filename=workbook_name)

reader = pd.read_excel(ethfile)
df = list(reader[etheader4])
series = reader[etheader4]
bottom = series.values[-1]
print (bottom)


#x = bottom
#try:
#	print ("It is a number " , float(x))
#except:
 #   bottom = bottom.replace(',', '')
  #  print (bottom)


today = date.today()
ethResponse = requests.get('https://api.etherscan.io/api?module=account&action=tokenbalance&contractaddress=' + contractAddr + '&address=' + walletAddr + '&tag=latest&apikey=' + ethAPIkey)
ethData = ethResponse.json() if ethResponse and ethResponse.status_code == 200 else None
ethtokenAmount = ethData["result"]
ethtokenAmount = float(ethtokenAmount)
ethtokenAmount = ethtokenAmount / 1000000000
print (ethtokenAmount)
# dd/mm/YY
d1 = today.strftime("%d/%m/%Y")  
d1 = today.strftime(datefor)       
coingecko = requests.get('https://api.coingecko.com/api/v3/simple/price?ids=' + tokenname + '&vs_currencies=' + tokencurr)
#pkData = coingecko.json() if coingecko and coingecko.status_code == 200 else None
cgData = coingecko.json()
ethtokenPrice = cgData[tokenname][tokencurr]
ethValue = ethtokenAmount * float(ethtokenPrice)
ethValue = str(ethValue)
ethOpening = float(bottom)
#print (ethOpening)
ethReflac =  ethtokenAmount - ethOpening
#print (ethReflac)
ethClosing = ethtokenAmount
ethReflacPer = ethReflac / ethOpening
#print (ethReflacPer)
ethslipagepercalint = float(ethslipageper)
ethslipagepercal = (100 - ethslipagepercalint) 
print (ethslipagepercal)
ethslipagepercal1 = (ethslipagepercal/100)
print (ethslipagepercal1)
ethSlipage = float(ethValue) * float(ethslipagepercal1)
print (ethSlipage)
ethReflacEarn = ethReflac * float(ethtokenPrice)

#building a pandas dataframe
dataf = [{  etheader1:d1, 
             etheader2:ethOpening, 
             etheader3:ethReflac,
             etheader4:ethClosing,
             etheader5: ethReflacPer, 
             etheader6: ethtokenPrice,
             etheader7:ethReflacEarn,
             etheader8:ethtextp,
             etheader9:ethValue,
             etheader10:ethSlipage}
            ]

# create a new workbook
workbook_name = ethfile
wb = load_workbook(workbook_name)
ws = wb[sfworksheet]

# Dictionarys are not in order by default
# Define a `list` of `keys` in desired order
fieldnames = [etheader1, etheader2, etheader3, etheader4, etheader5, etheader6,etheader7,etheader8,etheader9,etheader10]



# append data
# iterate `list` of `dict`
for list in dataf:
    # create a dataframe `value`
    # use the fieldnames in desired order as `key`
    values = (list[k] for k in fieldnames)
    # append the `generator values`



    
ws.append(values)

###############################################
### Formating of cells

for col in ws['B']:
    col.number_format = '#,##0.00'

for col in ws['C']:
    col.number_format = '#,##0.00'

for col in ws['D']:
    col.number_format = '#,##0.00'

for col in ws['E']:
    col.number_format = '0.0000000%'

for col in ws['F']:
    col.number_format = '#,##0.0000000000'

for col in ws['G']:
    col.number_format = '#,##0.0000'

for col in ws['I']:
    col.number_format = '#,##0.00'

for col in ws['I']:
    col.number_format = '#,##0.00'

for col in ws['J']:
    col.number_format = '#,##0.00'


##########################################################
    

wb.save(filename=workbook_name)

exit()



      
