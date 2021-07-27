# This script doesn't work if the closing balance column has thousand separator so only thing to fix
# Tip jar address if you appreciate this script 0x32D2da7308516e46D38EC638FE2144dE2250B7D3 Thank You 
#
#Big thanks again to Sam Brimhall
#After installing python you also need to install module requests and pandas to do this simple type in command prompt 
# following command 'python -m pip install requests' and python -m pip install pandas and 'python -m pip install openpyxl'


########################## INSERT YOUR INFO AS NEEDED BELOW ################################################

#Contract Address - the token you wish to read value from on Bscscan (set to Safemoon, DO NOT CHANGE)
contractAddr = '0x8076C74C5e3F5852037F31Ff0093Eeb8c8ADd8D3'

#Wallet Address - INSERT YOUR PUBLIC WALLET ADDRESS HERE
walletAddr = ''

#API Key for bscscan - MAKE AND ACCOUNT https://bscscan.com/register verify your account then go to https://bscscan.com/myapikey and generate a new API key. INSERT THAT API KEY HERE
bscAPIkey = ''

#Location and file name
sfmoonfile = 'c:/crypto/Safemoon-Tracker.xlsx'

#How the date gets recorded in Australia it is dd/mm/yyyy
datefor = "%d/%m/%Y"

#To get rid of an error it needs to have a start balance this figure is just used as movement between second row please use just whole numbers
# eg 1,000.34 please enter 1000 different regions type thousand seperator differently and it causing issues as I am learning I will update
startbal = '1'

#Name of the worksheet
sfworksheet = 'Safemoon Tracker'


#To get net sale figure default 12% slipage is used 
sfslipageper = '12'

#If you want to change the column header name
sfheader1 = 'Date'
sfheader2 = 'Opening Balance'
sfheader3 = 'Daily Gain'
sfheader4 = 'Closing Balance'
sfheader5 = '% Increase'
sfheader6 = 'Price'
sfheader7 = 'Earned Reflactions'
sfheader8 = 'Price From'
sfheader9 = 'Gross Value'
sfheader10 = 'Net Value'

# if you want to change text
sftextp = 'Pancakeswap price US'
  
############################################################################################################

###PROGRAM CODE BELOW - DO NOT TOUCH UNLESS YOU KNOW WHAT YOU ARE DOING###
import requests
import json
import time
import csv
import pandas as pd
import os
from datetime import date
from pandas import DataFrame
import locale
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.utils import get_column_letter

locale.setlocale(locale.LC_ALL, '')
col_names = [sfheader1,
             sfheader2,
             sfheader3,
             sfheader4,
             sfheader5,
             sfheader6,
             sfheader7,
             sfheader8,
             sfheader9,
             sfheader10,]

import os.path
#If command to check does the daily tracker file exists if it doesn't than in else create it and write
#colum headers
if os.path.isfile(sfmoonfile):
    print ("File exist")
else:
    print ("File not exist")
    df = pd.DataFrame([col_names])
# Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(sfmoonfile, engine='xlsxwriter')
# Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name=sfworksheet, index=False, header=False)  
# Close the Pandas Excel writer and output the Excel file.
    writer.save()
# open the workbook created to append a opening balance as engine xlsxwriter doesn't support appending rows
workbook_name = sfmoonfile
wb = load_workbook(workbook_name)
ws = wb[sfworksheet]
if ws.cell(row = 2, column=4).value == None:
    ws['D2'] = float(startbal)
    print("New Excel Spreadsheet")
    style_1 = Font(bold=True)
    for cell in ws["1:1"]:
      cell.font = style_1
    for col_range in range(1, 11):
        cell_title = ws.cell(1, col_range)
        cell_title.fill = PatternFill(start_color="00A79D", end_color="00A79D", fill_type="solid")
    for col_range in range(1, 11):
        cell_title = ws.cell(1, col_range)
        cell.font = style_1
else:
       print("Existing Worksheet")
#Save the spredsheet  
wb.save(filename=workbook_name)
#Open sheet to to pick up the closing value so it becames the opening value
reader = pd.read_excel(sfmoonfile)
df = list(reader[sfheader4])
series = reader[sfheader4]
bottom = series.values[-1]
print (bottom)


#x = bottom
#try:
#	print ("It is a number " , float(x))
#except:
 #   bottom = bottom.replace(',', '')
  #  print (bottom)


today = date.today()
bscResponse = requests.get('https://api.bscscan.com/api?module=account&action=tokenbalance&contractaddress=' + contractAddr + '&address=' + walletAddr + '&tag=latest&apikey=' + bscAPIkey)
bscData = bscResponse.json() if bscResponse and bscResponse.status_code == 200 else None
safemoonAmount = bscData["result"]
safemoonAmount = float(safemoonAmount)
safemoonAmount = safemoonAmount / 1000000000
# dd/mm/YY
#d1 = today.strftime("%d/%m/%Y")  
d1 = today.strftime(datefor)       
pkResponse = requests.get('https://api.pancakeswap.info/api/v2/tokens')
pkData = pkResponse.json() if pkResponse and pkResponse.status_code == 200 else None
safemoonPrice = pkData["data"][contractAddr]["price"]
safemoonPrice1 = float(safemoonPrice)
safemoonPrice2 = round(safemoonPrice1,10)
#print("SafeMoon Price: " + safemoonPrice)
sfValue = safemoonAmount * float(safemoonPrice)
sfValue = float(sfValue)
sfOpening = float(bottom)
#print (sfOpening)
sfReflac =  safemoonAmount - sfOpening
#print (sfReflac)
sfClosing = safemoonAmount
#time.sleep(10)
sfReflacPer = sfReflac / sfOpening
#print (sfReflacPer)
sfslipagepercalint = float(sfslipageper)
sfslipagepercal = (100 - sfslipagepercalint) 
#print (sfslipagepercal)
sfslipagepercal1 = (sfslipagepercal/100)
#print (sfslipagepercal1)
sfSlipage = float(sfValue) * float(sfslipagepercal1)
#print (sfSlipage)
sfReflacEarn = sfReflac * float(safemoonPrice)

#building a pandas dataframe
dataf = [{sfheader1:d1, 
             sfheader2:sfOpening, 
             sfheader3:sfReflac,
             sfheader4:sfClosing,
             sfheader5: sfReflacPer, 
             sfheader6:safemoonPrice2,
             sfheader7:sfReflacEarn,
             sfheader8:sftextp,
             sfheader9:sfValue,
             sfheader10:sfSlipage}
            ]

# create a new workbook
workbook_name = sfmoonfile
wb = load_workbook(workbook_name)
ws = wb[sfworksheet]

# Dictionarys are not in order by default
# Define a `list` of `keys` in desired order
fieldnames = [sfheader1, sfheader2, sfheader3, sfheader4, sfheader5, sfheader6,sfheader7,sfheader8,sfheader9,sfheader10]



# append data
# iterate `list` of `dict`
for list in dataf:
    # create a dataframe `value`
    # use the fieldnames in desired order as `key`
    values = (list[k] for k in fieldnames)
# append the `generator values` 
ws.append(values)

###############################################
# Formating of cells

for col in ws['B']:
    col.number_format = '#,##0.00'

for col in ws['C']:
    col.number_format = '#,##0.00'

for col in ws['D']:
    col.number_format = '#,##0.00'

for col in ws['E']:
    col.number_format = '0.0000000%'

for col in ws['F']:
    col.number_format = '#,##0.0000000'

for col in ws['G']:
    col.number_format = '#,##0.0000'

for col in ws['I']:
    col.number_format = '#,##0.00'

for col in ws['I']:
    col.number_format = '#,##0.00'

for col in ws['J']:
    col.number_format = '#,##0.00'
##########################################################
wb.save(filename=workbook_name)

exit()



      
